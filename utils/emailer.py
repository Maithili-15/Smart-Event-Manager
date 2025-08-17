# utils/emailer.py
from openpyxl import load_workbook
from email.message import EmailMessage
import smtplib
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv

load_dotenv()  # loads .env from project root

SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))
SMTP_EMAIL = os.getenv("SMTP_EMAIL")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
FROM_NAME = os.getenv("FROM_NAME", "Smart Event Manager")


def read_attendees_from_xlsx(path="attendees.xlsx"):
    """Return list of dicts: [{'name':..., 'email':..., 'event_id':...}, ...]"""
    attendees = []
    if not os.path.exists(path):
        raise FileNotFoundError(f"{path} not found")
    wb = load_workbook(filename=path)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_l = [h.lower() if h else "" for h in headers]
    # indices
    idx_name = headers_l.index("name") if "name" in headers_l else None
    idx_email = headers_l.index("email") if "email" in headers_l else None
    idx_eid = headers_l.index("event_id") if "event_id" in headers_l else None

    if idx_email is None:
        raise ValueError("attendees.xlsx must have an 'email' column")

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[idx_name] if idx_name is not None else None
        email = row[idx_email]
        event_id = row[idx_eid] if idx_eid is not None else None
        if email:
            attendees.append({
                "name": str(name).strip() if name else "",
                "email": str(email).strip(),
                "event_id": str(event_id).strip() if event_id else ""
            })
    return attendees


def build_reminder_text(event):
    when = f"{event.date} at {event.time}"
    text = f"Reminder: {event.name} on {when} at {event.location or 'N/A'}"
    html = f"""
    <html>
      <body>
        <h2>Event Reminder</h2>
        <p><b>{event.name}</b></p>
        <p>Date & Time: {when}</p>
        <p>Location: {event.location or 'N/A'}</p>
        <p>Type: {event.type}</p>
      </body>
    </html>
    """
    return text, html



def send_email(to_email, subject, text_body, html_body=None):
    """Send single email via SMTP (uses global SMTP_*)."""
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        raise RuntimeError("SMTP credentials are not configured in .env")

    msg = EmailMessage()
    msg["From"] = f"{FROM_NAME} <{SMTP_EMAIL}>"
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(text_body)
    if html_body:
        msg.add_alternative(html_body, subtype="html")

    # send
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
        smtp.ehlo()
        if SMTP_PORT in (587, 25):
            smtp.starttls()
            smtp.ehlo()
        smtp.login(SMTP_EMAIL, SMTP_PASSWORD)
        smtp.send_message(msg)


def send_reminders_for_events(events, attendees, dry_run=True):
    """
    events: list of Event objects
    attendees: list of {'name','email','event_id'} dicts
    dry_run: if True, print emails instead of sending
    """
    sent = []
    events_map = {str(e.id): e for e in events}   # map by string IDs

    for a in attendees:
        if a['event_id']:
            ev = events_map.get(a['event_id'])
            if not ev:
                continue
            text, html = build_reminder_text(ev)
            subject = f"Reminder: {ev.name} on {ev.date}"
            if dry_run:
                print(f"\n--- DRY RUN: To: {a['email']} ---\nSubject: {subject}\n{text}\n")
                sent.append((a['email'], subject))
            else:
                send_email(a['email'], subject, text, html)
                sent.append((a['email'], subject))
        else:
            # send all events if no specific event_id
            for ev in events:
                text, html = build_reminder_text(ev)
                subject = f"Reminder: {ev.name} on {ev.date}"
                if dry_run:
                    print(f"\n--- DRY RUN: To: {a['email']} ---\nSubject: {subject}\n{text}\n")
                    sent.append((a['email'], subject))
                else:
                    send_email(a['email'], subject, text, html)
                    sent.append((a['email'], subject))
    return sent


def get_upcoming_events(all_events, within_hours=None, specific_date=None):
    """
    all_events: list of event dicts
    within_hours: integer hours from now (e.g., 24)
    specific_date: string "DD-MM-YYYY" to filter exact date
    """
    results = []
    now = datetime.now()
    if within_hours is not None:
        deadline = now + timedelta(hours=within_hours)
        for e in all_events:
            try:
                dt = datetime.strptime(f"{e['date']} {e['time']}", "%d-%m-%Y %H:%M")
                if now <= dt <= deadline:
                    results.append(e)
            except Exception:
                continue
        return results
    if specific_date:
        for e in all_events:
            if e['date'] == specific_date:
                results.append(e)
        return results
    return []
